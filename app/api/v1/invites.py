import logging
import os
from fastapi import APIRouter, Depends, HTTPException, Security, UploadFile, status, Response
import openpyxl
from sqlmodel import Session, select
from app.api.v1.authentication import token_decode
from app.core.config import MEDIA_ADMIN_DIRECTORY
from app.models.invites import Invites
from app.schemas.invites import InvitesResponseSchema, InvitesCreateSchema, InvitesUpdateSchema
from app.core.database import get_session
from fastapi import Body

router = APIRouter()

# Configura el registrador
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


@router.post("/", status_code=status.HTTP_201_CREATED, response_model=InvitesResponseSchema)
def create_invite(
    invite: InvitesCreateSchema,
    db: Session = Depends(get_session),
    user=Security(token_decode, scopes=["admin"]),
):
    logger.info(f"Creating invite with data: {invite.model_dump()} {user}")

    invite_data = invite.model_dump()

    if db.exec(select(Invites).where(Invites.phone_number == invite.phone_number)).first():
        raise HTTPException(status_code=status.HTTP_409_CONFLICT,
                            detail="El invitado ya existe")

    db_invite = Invites(**invite_data)
    db.add(db_invite)
    db.commit()
    db.refresh(db_invite)
    logger.info(
        f"invite created with ID: {db_invite.id}, and values: {db_invite}")
    return db_invite


@router.get("/{invite_id}", response_model=InvitesResponseSchema)
def get_invite(
    invite_id: int,
    user=Depends(token_decode),
    db: Session = Depends(get_session)
):
    logger.info(f"Getting invite with ID: {invite_id}")
    statement = select(Invites).where(
        Invites.id == invite_id,
    )
    result = db.exec(statement)
    db_invite = result.first()
    logger.info(f"Getting invite data: {db_invite}")
    if not db_invite:
        logger.error(f"invite with ID: {invite_id} not found")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="invite not found"
        )
    return db_invite


@router.get("/", response_model=list[InvitesResponseSchema])
def get_invites(
    user=Security(token_decode, scopes=["admin"]),
    db: Session = Depends(get_session)
):

    if not user.get("is_admin", False):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, detail="No autorizado")

    statement = select(Invites)
    result = db.exec(statement)
    db_invites = result.all()
    return db_invites


@router.patch("/{invite_id}", response_model=InvitesResponseSchema)
def update_invite(
    invite_id: int,
    invite: InvitesUpdateSchema,
    user=Security(token_decode, scopes=["admin"]),
    db: Session = Depends(get_session)
):
    statement = select(Invites).where(Invites.id == invite_id)
    result = db.exec(statement)
    db_invite = result.first()
    if not db_invite:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="invite not found"
        )
    update_data = invite.model_dump(exclude_unset=True)
    db_invite.sqlmodel_update(update_data)
    db.commit()
    db.refresh(db_invite)
    return db_invite


@router.delete("/{invite_id}")
def delete_invite(
    invite_id: int,
    user=Security(token_decode, scopes=["admin"]),
    db: Session = Depends(get_session)
):
    statement = select(Invites).where(Invites.id == invite_id)
    result = db.exec(statement)
    db_invite = result.first()
    if not db_invite:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="invite not found"
        )
    db.delete(db_invite)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post("/upload_invites/")
async def create_upload_file(
    file: UploadFile,
    user=Security(token_decode, scopes=["admin"]),
    db: Session = Depends(get_session)
):
    print(file.filename)

    if "officedocument" not in file.content_type:
        raise HTTPException(status_code=400, detail="El archivo subido no es el esperado (xlsx)")

    file_path = os.path.join(MEDIA_ADMIN_DIRECTORY, file.filename)

    # Guardar el archivo en el servidor
    with open(file_path, "wb") as buffer:
        buffer.write(file.file.read())

    wb = openpyxl.load_workbook(os.path.join(MEDIA_ADMIN_DIRECTORY, file.filename))
    ws = wb.active
    print('Total number of rows: '+str(ws.max_row) + '. And total number of columns: '+str(ws.max_column))

    # users_list = []

    # for value in ws.iter_rows(min_row=1, values_only=True):
    #     users_list.append(value)

    for row in ws.iter_rows(min_row=2, max_col=6):
        invite = []

        for index, cell in enumerate(row):
            invite.append(cell.value)

        new_invite = Invites(name=invite[0], last_name=invite[1], gender=invite[2],
                             birth_date=invite[3], phone_number=invite[4], address=invite[5])
        db.add(new_invite)
    db.commit()

    return {"filename": file.filename}
